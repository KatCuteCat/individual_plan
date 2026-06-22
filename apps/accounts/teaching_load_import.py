"""
Парсер выгрузки учебной нагрузки из 1С.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Optional

import openpyxl

from apps.core.models import TeachingActivityType

# Структуры данных


@dataclass
class ParsedItem:
    """
    Одна разобранная строка учебной нагрузки.
    """
    row: int                                # номер строки в Excel (для отчёта об ошибках)
    discipline: str
    semester: int                           # номер семестра / сессии
    activity_type: TeachingActivityType
    hours: Decimal
    group_number: str = ''                  # колонка «Факультет, курс, группа»
    cycle: str = ''                         # цикл дисциплины (напр. «Б1.О.23»)
    students_count: Optional[int] = None    # число студентов
    period_label: str = ''  # исходный текст периода контроля (col5 / col3)
    row_num: Optional[int] = None  # порядковый номер строки из col1


# Синонимы видов нагрузки 1С → код TeachingActivityType


ACTIVITY_TYPE_SYNONYMS = {
    # Лекции
    'лекционные занятия': 'LECTURE',
    'лекции': 'LECTURE',

    # Лабораторки
    'лабораторные занятия': 'LAB',
    'лабораторные работы': 'LAB',

    # Практики (аудиторные практические занятия)
    'практические занятия': 'PRACTICE_LESSON',

    # Консультации
    'консультация': 'CONSULTATION',
    'консультации': 'CONSULTATION',
    'консультации к лекционным занятиям, экзаменам': 'CONSULTATION',
    'консультации к практикам': 'CONSULTATION',

    # Самостоятельная работа под руководством преподавателя (СРП)
    'самостоятельная работа под руководством преподавателя': 'SUPERVISED_SELF_STUDY',

    # Контроль
    'зачет': 'CREDIT',
    'дифференцированный зачет': 'GRADED_CREDIT',
    'дифзачет': 'GRADED_CREDIT',
    'экзамен': 'EXAM',

    # Курсовые
    'курсовая работа': 'COURSE_WORK',
    'курсовой проект': 'COURSE_PROJECT',

    # ВКР
    'руководство вкр': 'THESIS_SUPERVISION',
    'руководство выпускной квалификационной работой': 'THESIS_SUPERVISION',
    'защита вкр': 'THESIS_DEFENSE',

    # Практики (учебные / производственные / технологические)
    'прочая практика': 'OTHER_PRACTICE',

    # Аспиранты
    'руководство аспирантами': 'POSTGRADUATE_SUPERVISION',
}



# Маппинг периода контроля → номер семестра

SEMESTER_BY_WORD = {
    'первый семестр': 1,
    'второй семестр': 2,
    'третий семестр': 3,
    'четвертый семестр': 4,
    'пятый семестр': 5,
    'шестой семестр': 6,
    'седьмой семестр': 7,
    'восьмой семестр': 8,
}

_SESSION_RE = re.compile(
    r'^\s*(\d+)\s+сессия\s*\(\s*(зимняя|летняя)\s*\)\s*$',
    re.IGNORECASE,
)



def _normalize(value) -> str:
    """Нижний регистр, ё→е, схлопывание пробелов, strip."""
    if value is None:
        return ''
    s = str(value).lower().replace('ё', 'е')
    s = re.sub(r'\s+', ' ', s)
    return s.strip()


def _try_int(value) -> Optional[int]:

    if value is None:
        return None
    s = str(value).strip().replace(',', '.').replace(' ', '')
    if not s:
        return None
    try:
        f = float(s)
        n = int(f)
        if f == n and n > 0:
            return n
        return None
    except (ValueError, TypeError):
        return None


def _parse_hours(value) -> Optional[Decimal]:
    """
    Преобразует значение ячейки «План» в Decimal.
    Поддерживает числа (int/float), строки с запятой («3,91»)
    и строки с точкой («3.91»). Возвращает None для пустых/невалидных.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None
    s = str(value).strip()
    if not s:
        return None
    s = s.replace(',', '.').replace(' ', '')
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _parse_period_to_semester(period_value) -> Optional[int]:

    norm = _normalize(period_value)
    if not norm:
        return None

    if norm in SEMESTER_BY_WORD:
        return SEMESTER_BY_WORD[norm]

    m = _SESSION_RE.match(norm)
    if m:
        n = int(m.group(1))
        kind = m.group(2)
        if 1 <= n <= 8:
            return n
        return 7 if kind == 'зимняя' else 8

    return None


def _resolve_activity_type(
    activity_raw,
    row_idx: int,
    discipline: str,
    activity_types_by_code: dict,
    errors: list[str],
) -> Optional[TeachingActivityType]:

    activity_norm = _normalize(activity_raw)
    code = ACTIVITY_TYPE_SYNONYMS.get(activity_norm)
    if code is None:
        errors.append(
            f'Строка {row_idx}: неизвестный вид нагрузки «{activity_raw}» '
            f'(дисциплина: «{discipline}»). '
            f'Добавьте синоним в teaching_load_import.py или новый вид в справочник.'
        )
        return None

    activity_type = activity_types_by_code.get(code)
    if activity_type is None:
        errors.append(
            f'Строка {row_idx}: вид нагрузки с кодом «{code}» не найден в справочнике '
            f'TeachingActivityType. Загрузите fixture teaching_activity_types.'
        )
        return None

    return activity_type



# Детектор шапки таблицы 1.2


def _find_header_row(ws) -> Optional[int]:
    """
    Ищет строку-шапку таблицы 1.2 «Занятия по учебным дисциплинам».
    Условие совпадения:
      - col 2 == «Дисциплина»
      - col 7 == «Вид учебной нагрузки»
    Сканирует первые 200 строк. Возвращает номер строки или None.
    """
    limit = min(ws.max_row or 0, 200)
    for row in range(1, limit + 1):
        c2 = _normalize(ws.cell(row, 2).value)
        c7 = _normalize(ws.cell(row, 7).value)
        if c2 == 'дисциплина' and c7 == 'вид учебной нагрузки':
            return row
    return None


# Парсер полного плана (раздел 1.2)

def _parse_full_plan(
    ws,
    header_row: int,
    activity_types_by_code: dict,
) -> tuple[list[ParsedItem], list[str], list[str]]:
    """
    Парсит раздел 1.2 полного индивидуального плана.

    Колонки данных:
      col 2  — дисциплина
      col 5  — период контроля (семестр)
      col 7  — вид учебной нагрузки
      col 9  — цикл (напр. «Б1.О.23»)
      col 11 — факультет, форма обучения, курс, группа
      col 13 — число студентов
      col 14 — часы (план)
      col 16 — часы (факт из 1С) — НЕ читаем
    """
    items: list[ParsedItem] = []
    errors: list[str] = []
    warnings: list[str] = []


    data_start = None
    search_limit = min(header_row + 15, (ws.max_row or 0) + 1)
    for row in range(header_row + 1, search_limit):
        c1 = ws.cell(row, 1).value
        c2 = ws.cell(row, 2).value
        n = _try_int(c1)
        if n is not None and n > 0 and _try_int(c2) is None:
            data_start = row
            break

    if data_start is None:
        errors.append(
            f'Не удалось найти начало данных после шапки таблицы 1.2 '
            f'(шапка на строке {header_row}).'
        )
        return items, errors, warnings

    # Читаем строки данных
    max_row = min(ws.max_row or 0, data_start + 500)

    for row_idx in range(data_start, max_row + 1):
        c1 = ws.cell(row_idx, 1).value
        c2 = ws.cell(row_idx, 2).value
        c7 = ws.cell(row_idx, 7).value

        # Стоп-маркеры
        # 1. Полностью пустая строка (col1, col2, col7 все None)
        if c1 is None and c2 is None and c7 is None:
            break
        # 2. «Итого» в col2
        if c2 is not None and 'итого' in _normalize(c2):
            break
        # 3. Нечисловой текст в col1 (заголовок следующего раздела)
        if c1 is not None and _try_int(c1) is None:
            break

        # Читаем ячейки
        discipline_raw = c2
        period_raw = ws.cell(row_idx, 5).value
        activity_raw = c7
        cycle_raw = ws.cell(row_idx, 9).value
        group_raw = ws.cell(row_idx, 11).value
        students_raw = ws.cell(row_idx, 13).value
        hours_raw = ws.cell(row_idx, 14).value

        discipline = str(discipline_raw).strip() if discipline_raw else ''
        group_number = str(group_raw).strip() if group_raw else ''
        cycle = str(cycle_raw).strip() if cycle_raw else ''
        students_count = _try_int(students_raw)

        # Часы (план)
        hours = _parse_hours(hours_raw)
        if hours is None or hours <= 0:
            warnings.append(
                f'Строка {row_idx}: пропущена — план не указан '
                f'(дисциплина: «{discipline}», вид: «{activity_raw}»).'
            )
            continue

        # Дисциплина
        if not discipline:
            errors.append(f'Строка {row_idx}: не указана дисциплина.')
            continue

        # Семестр
        semester = _parse_period_to_semester(period_raw)
        if semester is None:
            errors.append(
                f'Строка {row_idx}: не удалось распознать период контроля '
                f'«{period_raw}» (дисциплина: «{discipline}»).'
            )
            continue

        # Вид нагрузки
        activity_type = _resolve_activity_type(
            activity_raw, row_idx, discipline, activity_types_by_code, errors,
        )
        if activity_type is None:
            continue

        items.append(ParsedItem(
            row=row_idx,
            discipline=discipline,
            semester=semester,
            activity_type=activity_type,
            hours=hours,
            group_number=group_number,
            cycle=cycle,
            students_count=students_count,
            period_label=str(period_raw).strip() if period_raw else '',
            row_num=_try_int(c1),
        ))

    if not items and not errors:
        errors.append(
            'В таблице 1.2 не найдено ни одной строки учебной нагрузки.'
        )

    return items, errors, warnings



# Парсер короткого отчёта


def _parse_short_format(
    ws,
    activity_types_by_code: dict,
) -> tuple[list[ParsedItem], list[str], list[str]]:

    items: list[ParsedItem] = []
    errors: list[str] = []
    warnings: list[str] = []

    max_row = min(ws.max_row or 0, 1000)
    stop_row = None

    for row_idx in range(10, max_row + 1):
        a_value = ws.cell(row=row_idx, column=1).value

        if isinstance(a_value, str) and 'итого' in _normalize(a_value):
            stop_row = row_idx
            break

        # Пропускаем пустые строки
        if all(
            ws.cell(row=row_idx, column=c).value is None
            or not str(ws.cell(row=row_idx, column=c).value).strip()
            for c in range(1, 14)
        ):
            continue

        discipline_raw = ws.cell(row=row_idx, column=2).value
        period_raw = ws.cell(row=row_idx, column=3).value
        activity_raw = ws.cell(row=row_idx, column=4).value
        group_raw = ws.cell(row=row_idx, column=5).value
        plan_raw = ws.cell(row=row_idx, column=6).value

        discipline = str(discipline_raw).strip() if discipline_raw is not None else ''
        group_number = str(group_raw).strip() if group_raw is not None else ''

        # Часы
        hours = _parse_hours(plan_raw)
        if hours is None or hours <= 0:
            warnings.append(
                f'Строка {row_idx}: пропущена — план не указан '
                f'(дисциплина: «{discipline}», вид: «{activity_raw}»).'
            )
            continue

        # Дисциплина
        if not discipline:
            errors.append(f'Строка {row_idx}: не указана дисциплина.')
            continue

        # Семестр
        semester = _parse_period_to_semester(period_raw)
        if semester is None:
            errors.append(
                f'Строка {row_idx}: не удалось распознать период контроля '
                f'«{period_raw}» (дисциплина: «{discipline}»).'
            )
            continue

        # Вид нагрузки
        activity_type = _resolve_activity_type(
            activity_raw, row_idx, discipline, activity_types_by_code, errors,
        )
        if activity_type is None:
            continue

        items.append(ParsedItem(
            row=row_idx,
            discipline=discipline,
            semester=semester,
            activity_type=activity_type,
            hours=hours,
            group_number=group_number,
            period_label=str(period_raw).strip() if period_raw else '',
        ))

    if stop_row is None:
        warnings.append(
            'В файле не найдена строка «Итого учебная нагрузка:» — '
            'возможно, выгрузка обрезана. Проверьте файл.'
        )

    if not items and not errors:
        errors.append('В файле не найдено ни одной строки учебной нагрузки.')

    return items, errors, warnings

def parse_teaching_load_xlsx(file) -> tuple[list[ParsedItem], list[str], list[str]]:
    """
    Парсит Excel-выгрузку учебной нагрузки из 1С.
    """
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception as e:
        return [], [f'Не удалось открыть Excel-файл: {e}'], []

    ws = wb.active
    if ws is None:
        return [], ['В файле нет активного листа.'], []

    # Загружаем справочник видов нагрузки (один запрос)
    activity_types_by_code = {
        at.code: at for at in TeachingActivityType.objects.all()
    }

    # 1. Ищем шапку таблицы 1.2 в любом месте файла
    header_row = _find_header_row(ws)
    if header_row is not None:
        return _parse_full_plan(ws, header_row, activity_types_by_code)

    # 2. Fallback: короткий отчёт (A1 = «Отчет»)
    a1 = _normalize(ws.cell(1, 1).value)
    if a1 == 'отчет':
        return _parse_short_format(ws, activity_types_by_code)

    # 3. Ничего не подошло
    return [], [
        'Не удалось определить формат файла. '
        'Файл должен содержать либо таблицу с колонками '
        '«Дисциплина» и «Вид учебной нагрузки» (полный план), '
        'либо начинаться со слова «Отчет» (короткий отчёт).'
    ], []